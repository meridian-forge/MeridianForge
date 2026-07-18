from pathlib import Path

from openpyxl import load_workbook

from meridianforge.intake.extracted_data import ExtractedData
from meridianforge.intake.extractors.base import Extractor


class ExcelExtractor(Extractor):

    def extract(self, file_path: Path) -> ExtractedData:

        workbook = load_workbook(
            file_path,
            data_only=True,
        )

        fields: dict[str, str] = {}

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(value)
                    for value in row
                    if value is not None
                ]

                if len(values) >= 2:
                    fields[values[0]] = values[1]

        return ExtractedData(
            source_file=file_path.name,
            fields=fields,
        )
