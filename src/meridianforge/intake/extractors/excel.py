from pathlib import Path

from openpyxl import load_workbook

from meridianforge.intake.extracted_data import ExtractedData
from meridianforge.intake.extractors.base import Extractor


class ExcelExtractor(Extractor):
    """
    Excel extractor.

    Supports:

    1. Tabular workbooks
       address | city | state | purchase_price | ...

    2. Key/value workbooks
       Field | Value
    """

    def extract(
        self,
        file_path: Path,
    ) -> ExtractedData:

        workbook = load_workbook(
            file_path,
            data_only=True,
        )

        fields: dict[str, object] = {}

        for sheet in workbook.worksheets:

            rows = list(
                sheet.iter_rows(
                    values_only=True,
                )
            )

            if not rows:
                continue

            cleaned_rows = [
                [
                    value
                    for value in row
                    if value is not None
                ]
                for row in rows
            ]

            cleaned_rows = [
                row
                for row in cleaned_rows
                if len(row) > 0
            ]

            if not cleaned_rows:
                continue

            # Detect legacy key/value format first.
            # Example:
            # purchase_price | 250000
            # rent            | 2500
            key_value_rows = True

            for row in cleaned_rows:
                if len(row) < 2:
                    key_value_rows = False
                    break

            if key_value_rows:
                for row in cleaned_rows:
                    fields[str(row[0]).strip()] = row[1]

                break

            # Detect tabular format.
            # Example:
            # address | price | rent
            # Main St | 250000 | 2500
            if len(cleaned_rows) >= 2:

                headers = [
                    str(value).strip()
                    for value in cleaned_rows[0]
                ]

                first_row = cleaned_rows[1]

                if headers and len(first_row) >= len(headers):

                    for header, value in zip(
                        headers,
                        first_row,
                        strict=False,
                    ):
                        fields[header] = value

                    break

        return ExtractedData(
            source_file=file_path.name,
            fields=fields,
        )
