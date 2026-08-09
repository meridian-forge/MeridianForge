"""
Inventory workbook extractor.

MF-512.4.3

Parses turnkey inventory spreadsheets (such as the RTR inventory workbook)
and converts each valid property row into a structured inventory record that
can later be normalized and underwritten by the acquisition pipeline.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class InventoryWorkbookRecord:
    state: str
    asset_type: str
    address: str
    price: int
    roi: float | None
    cash_flow: int | None
    initial_cash: int | None
    beds: float | None
    baths: float | None
    year_built: int | None
    seller_incentives: str | None
    source_file: Path


class InventoryWorkbookExtractor:
    """
    Extract structured inventory records from an inventory workbook.

    The extractor intentionally ignores marketing / banner rows and only
    returns rows containing a numeric purchase price.
    """

    SHEET_NAME = "RTR Available Inventory"

    @classmethod
    def extract(cls, workbook_path: Path) -> list[InventoryWorkbookRecord]:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)

        if cls.SHEET_NAME in wb.sheetnames:
            ws = wb[cls.SHEET_NAME]
        else:
            ws = wb[wb.sheetnames[0]]

        records: list[InventoryWorkbookRecord] = []

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue

            state = row[0]
            asset_type = row[1]
            address = row[2]
            price = row[5]
            roi = row[6]
            cash_flow = row[7]
            initial_cash = row[8]
            beds = row[9]
            baths = row[10]
            year_built = row[11]
            seller_incentives = row[12]

            # Ignore marketing rows / blank rows
            if price is None:
                continue

            try:
                price_value = int(float(price))
            except (TypeError, ValueError):
                continue

            records.append(
                InventoryWorkbookRecord(
                    state=str(state).strip() if state else "",
                    asset_type=str(asset_type).strip() if asset_type else "",
                    address=str(address).strip() if address else "",
                    price=price_value,
                    roi=float(roi) if roi is not None else None,
                    cash_flow=int(float(cash_flow)) if cash_flow is not None else None,
                    initial_cash=(
                        int(float(initial_cash)) if initial_cash is not None else None
                    ),
                    beds=float(beds) if beds is not None else None,
                    baths=float(baths) if baths is not None else None,
                    year_built=(
                        int(float(year_built)) if year_built is not None else None
                    ),
                    seller_incentives=(
                        str(seller_incentives).strip() if seller_incentives else None
                    ),
                    source_file=workbook_path,
                )
            )

        return records
