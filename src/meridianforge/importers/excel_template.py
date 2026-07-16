"""
Excel template generator.

Creates Meridian Forge property import workbook.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelTemplateGenerator:
    """
    Generates investor-facing Excel templates.
    """

    HEADERS = [
        "purchase_price",
        "monthly_rent",
        "property_tax",
        "insurance",
        "hoa",
        "state",
        "city",
        "zip_code",
        "provider",
    ]

    @staticmethod
    def generate(
        output_path: str,
    ) -> Path:
        """
        Generate Meridian Forge property template.
        """

        path = Path(output_path)

        path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()

        properties = workbook.active
        properties.title = "Properties"

        properties.append(ExcelTemplateGenerator.HEADERS)

        properties.append(
            [
                250000,
                2200,
                3000,
                1500,
                0,
                "FL",
                "Jacksonville",
                "32210",
                "Example Provider",
            ]
        )

        instructions = workbook.create_sheet("Instructions")

        instructions.append(
            [
                "Field",
                "Description",
            ]
        )

        instructions_data = [
            (
                "purchase_price",
                "Property acquisition price",
            ),
            (
                "monthly_rent",
                "Expected monthly rental income",
            ),
            (
                "property_tax",
                "Annual property taxes",
            ),
            (
                "insurance",
                "Annual insurance cost",
            ),
            (
                "hoa",
                "Annual HOA cost",
            ),
            (
                "state",
                "Property state",
            ),
            (
                "city",
                "Property city",
            ),
            (
                "zip_code",
                "Property ZIP code",
            ),
            (
                "provider",
                "Source/provider name",
            ),
        ]

        for row in instructions_data:
            instructions.append(row)

        mapping = workbook.create_sheet("Mapping")

        mapping.append(
            [
                "Common Input",
                "Meridian Forge Field",
            ]
        )

        mapping_data = [
            ("Price", "purchase_price"),
            ("Purchase Cost", "purchase_price"),
            ("Rent", "monthly_rent"),
            ("Monthly Income", "monthly_rent"),
            ("Taxes", "property_tax"),
            ("Insurance", "insurance"),
        ]

        for row in mapping_data:
            mapping.append(row)

        for sheet in workbook:
            for column in sheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value else 0 for cell in column
                )

                sheet.column_dimensions[get_column_letter(column[0].column)].width = (
                    max_length + 5
                )

        workbook.save(path)

        return path
